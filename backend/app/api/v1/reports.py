from datetime import date, datetime, timedelta, timezone
from uuid import UUID
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, desc, or_, case
import io

from app.db.session import get_db
from app.core.dependencies import require_super_admin
from app.models.user import User
from app.models.subscription import Subscription, SubscriptionStatus, SubscriptionItem
from app.models.payment import Payment, PaymentStatus
from app.models.subscription_delivery import SubscriptionDelivery, DeliveryStatus
from app.models.delivery_assignment import DeliveryAssignment, AssignmentStatus
from app.models.delivery_partner import DeliveryPartner
from app.models.customer import Customer
from app.models.product import Product, ProductCategory
from app.models.fruit import Fruit, FruitOrder, FruitOrderItem, FruitAvailability, FruitOrderStatus, FruitPaymentStatus

router = APIRouter(prefix="/reports", tags=["Reports & Analytics"])

def parse_dates(start_date: str | None, end_date: str | None):
    try:
        sd = datetime.strptime(start_date.strip(), "%Y-%m-%d").date() if (start_date and start_date.strip() and start_date.strip() != "null") else date(2020, 1, 1)
        ed = datetime.strptime(end_date.strip(), "%Y-%m-%d").date() if (end_date and end_date.strip() and end_date.strip() != "null") else (date.today() + timedelta(days=1))
        return sd, ed
    except Exception:
        return date(2020, 1, 1), (date.today() + timedelta(days=1))

@router.get("/overview")
async def get_overview(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    
    # 1. Total Revenue
    pkg_rev = await db.scalar(select(func.coalesce(func.sum(Payment.amount), 0)).where(and_(Payment.status == PaymentStatus.SUCCESS, func.date(Payment.paid_at) >= sd, func.date(Payment.paid_at) <= ed)))
    frt_rev = await db.scalar(select(func.coalesce(func.sum(FruitOrder.total_amount), 0)).where(and_(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS, func.date(FruitOrder.paid_at) >= sd, func.date(FruitOrder.paid_at) <= ed)))
    total_revenue = float(pkg_rev or 0) + float(frt_rev or 0)
    
    # 2. Total Orders (Subscriptions + Fruit Orders)
    pkg_orders = await db.scalar(select(func.count(Subscription.id)).where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed, Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED]))))
    frt_orders = await db.scalar(select(func.count(FruitOrder.id)).where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed, FruitOrder.payment_status == FruitPaymentStatus.SUCCESS)))
    total_orders = (pkg_orders or 0) + (frt_orders or 0)
    
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # 3. Total Customers
    total_customers = await db.scalar(select(func.count(Customer.id)))
    
    # 4. Total Deliveries & Pending (for the period)
    total_deliv = await db.scalar(select(func.count(SubscriptionDelivery.id)).where(and_(func.date(SubscriptionDelivery.scheduled_date) >= sd, func.date(SubscriptionDelivery.scheduled_date) <= ed)))
    pending_deliv = await db.scalar(select(func.count(SubscriptionDelivery.id)).where(and_(func.date(SubscriptionDelivery.scheduled_date) >= sd, func.date(SubscriptionDelivery.scheduled_date) <= ed, SubscriptionDelivery.status == DeliveryStatus.PENDING)))
    
    # 5. Cancelled Orders
    canc_pkg = await db.scalar(select(func.count(Subscription.id)).where(and_(Subscription.status == SubscriptionStatus.CANCELLED, func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed)))
    canc_frt = await db.scalar(select(func.count(FruitOrder.id)).where(and_(FruitOrder.order_status == FruitOrderStatus.CANCELLED, func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed)))
    cancelled_orders = (canc_pkg or 0) + (canc_frt or 0)
    
    # 6. Active Subscriptions
    active_subs = await db.scalar(select(func.count(Subscription.id)).where(Subscription.status == SubscriptionStatus.ACTIVE))

    return {
        "total_revenue": total_revenue,
        "package_revenue": float(pkg_rev or 0),
        "fruit_revenue": float(frt_rev or 0),
        "total_orders": total_orders,
        "package_orders_count": pkg_orders or 0,
        "fruit_orders_count": frt_orders or 0,
        "total_customers": total_customers or 0,
        "total_deliveries": total_deliv or 0,
        "pending_deliveries": pending_deliv or 0,
        "cancelled_orders": cancelled_orders,
        "avg_order_value": avg_order_value,
        "active_subscriptions": active_subs or 0
    }

@router.get("/sales")
async def get_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    # Return daily revenue grouped by date
    pkg_stmt = select(func.date(Payment.paid_at).label('d'), func.sum(Payment.amount)).where(and_(Payment.status == PaymentStatus.SUCCESS, func.date(Payment.paid_at) >= sd, func.date(Payment.paid_at) <= ed)).group_by(func.date(Payment.paid_at))
    frt_stmt = select(func.date(FruitOrder.paid_at).label('d'), func.sum(FruitOrder.total_amount)).where(and_(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS, func.date(FruitOrder.paid_at) >= sd, func.date(FruitOrder.paid_at) <= ed)).group_by(func.date(FruitOrder.paid_at))
    
    pkg_res = await db.execute(pkg_stmt)
    frt_res = await db.execute(frt_stmt)
    
    sales_dict = {}
    for d, amt in pkg_res.all():
        if d:
            ds = str(d)
            sales_dict[ds] = sales_dict.get(ds, 0) + float(amt)
    for d, amt in frt_res.all():
        if d:
            ds = str(d)
            sales_dict[ds] = sales_dict.get(ds, 0) + float(amt)
        
    daily_sales = [{"date": k, "revenue": v} for k, v in sorted(sales_dict.items())]
    return {"daily_sales": daily_sales}

@router.get("/packages")
async def get_packages(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    stmt = (
        select(Product.name, func.count(Subscription.id).label('c'), func.sum(Subscription.total_amount))
        .join(Subscription, Subscription.product_id == Product.id)
        .where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed, Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED])))
        .group_by(Product.id, Product.name)
        .order_by(desc('c'))
    )
    res = await db.execute(stmt)
    items = []
    for row in res.all():
        items.append({"name": row[0], "orders": row[1], "revenue": float(row[2] or 0)})
    return {
        "top_selling": items[:5],
        "least_selling": items[-5:] if len(items) > 5 else list(reversed(items)),
        "all": items
    }

@router.get("/fruits")
async def get_fruits(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    stmt = (
        select(Fruit.name, func.sum(FruitOrderItem.quantity_kg).label('q'), func.sum(FruitOrderItem.price_per_kg * FruitOrderItem.quantity_kg).label('rev'))
        .join(FruitOrderItem, FruitOrderItem.fruit_id == Fruit.id)
        .join(FruitOrder, FruitOrder.id == FruitOrderItem.order_id)
        .where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed, FruitOrder.payment_status == FruitPaymentStatus.SUCCESS))
        .group_by(Fruit.id, Fruit.name)
        .order_by(desc('q'))
    )
    res = await db.execute(stmt)
    items = []
    for row in res.all():
        items.append({"name": row[0], "quantity": int(row[1] or 0), "revenue": float(row[2] or 0)})
        
    return {
        "top_selling": items[:5],
        "least_selling": items[-5:] if len(items) > 5 else list(reversed(items)),
        "all": items
    }

@router.get("/customers")
async def get_customers(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    total = await db.scalar(select(func.count(Customer.id)))
    new_c = await db.scalar(select(func.count(Customer.id)).where(and_(func.date(Customer.created_at) >= sd, func.date(Customer.created_at) <= ed)))
    return {
        "total": total,
        "new": new_c,
        "active": total, 
        "inactive": 0
    }

@router.get("/deliveries")
async def get_deliveries(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    stmt = select(SubscriptionDelivery.status, func.count(SubscriptionDelivery.id)).where(and_(func.date(SubscriptionDelivery.scheduled_date) >= sd, func.date(SubscriptionDelivery.scheduled_date) <= ed)).group_by(SubscriptionDelivery.status)
    res = await db.execute(stmt)
    status_counts = {str(k.value if hasattr(k, 'value') else k): v for k, v in res.all()}
    
    return {
        "status_breakdown": status_counts
    }

@router.get("/orders")
async def get_orders(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    pkg_stmt = select(Subscription.status, func.count(Subscription.id)).where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed)).group_by(Subscription.status)
    pkg_res = await db.execute(pkg_stmt)
    pkg_counts = {str(k.value if hasattr(k, 'value') else k): v for k, v in pkg_res.all()}
    
    frt_stmt = select(FruitOrder.order_status, func.count(FruitOrder.id)).where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed)).group_by(FruitOrder.order_status)
    frt_res = await db.execute(frt_stmt)
    frt_counts = {str(k.value if hasattr(k, 'value') else k): v for k, v in frt_res.all()}
    
    return {
        "package_orders": pkg_counts,
        "fruit_orders": frt_counts,
    }

@router.get("/payments")
async def get_payments(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    sd, ed = parse_dates(start_date, end_date)
    pkg_stmt = select(Payment.status, func.count(Payment.id), func.sum(Payment.amount)).where(and_(func.date(Payment.created_at) >= sd, func.date(Payment.created_at) <= ed)).group_by(Payment.status)
    pkg_res = await db.execute(pkg_stmt)
    
    counts = {}
    total_collected = 0.0
    for st, c, amt in pkg_res.all():
        if st is not None:
            s = str(st.value if hasattr(st, 'value') else st)
            counts[s] = counts.get(s, 0) + c
            if s == 'success':
                total_collected += float(amt or 0)
            
    return {
        "status_counts": counts,
        "total_collected": total_collected
    }

@router.get("/export/excel")
async def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_super_admin)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    sd, ed = parse_dates(start_date, end_date)
    
    # Query summary metrics for the given date range
    pkg_rev = await db.scalar(select(func.coalesce(func.sum(Payment.amount), 0)).where(and_(Payment.status == PaymentStatus.SUCCESS, func.date(Payment.paid_at) >= sd, func.date(Payment.paid_at) <= ed)))
    frt_rev = await db.scalar(select(func.coalesce(func.sum(FruitOrder.total_amount), 0)).where(and_(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS, func.date(FruitOrder.paid_at) >= sd, func.date(FruitOrder.paid_at) <= ed)))
    total_revenue = float(pkg_rev or 0) + float(frt_rev or 0)
    
    pkg_orders = await db.scalar(select(func.count(Subscription.id)).where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed, Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED]))))
    frt_orders = await db.scalar(select(func.count(FruitOrder.id)).where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed, FruitOrder.payment_status == FruitPaymentStatus.SUCCESS)))
    total_orders = (pkg_orders or 0) + (frt_orders or 0)
    
    active_subs = await db.scalar(select(func.count(Subscription.id)).where(Subscription.status == SubscriptionStatus.ACTIVE))
    pending_deliv = await db.scalar(select(func.count(SubscriptionDelivery.id)).where(and_(func.date(SubscriptionDelivery.scheduled_date) >= sd, func.date(SubscriptionDelivery.scheduled_date) <= ed, SubscriptionDelivery.status == DeliveryStatus.PENDING)))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1B5E20")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    # Title Banner
    ws.merge_cells("A1:D1")
    ws["A1"] = "HEALTHY HOME FOODS - BUSINESS REPORT"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = f"Period: {sd.strftime('%d %b %Y')} to {ed.strftime('%d %b %Y')}"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["C2"] = f"Generated On: {date.today().strftime('%d %b %Y')}"
    ws["C2"].font = Font(italic=True, size=10, color="555555")
    
    ws.append([])
    
    # Section 1: Business Overview
    ws.append(["EXECUTIVE METRICS", "", "", ""])
    curr_row = ws.max_row
    ws.merge_cells(f"A{curr_row}:D{curr_row}")
    ws[f"A{curr_row}"].font = section_font
    ws[f"A{curr_row}"].fill = section_fill
    
    metrics = [
        ("Total Revenue", f"Rs. {total_revenue:,.2f}"),
        ("Package Revenue", f"Rs. {float(pkg_rev or 0):,.2f}"),
        ("Grocery Revenue", f"Rs. {float(frt_rev or 0):,.2f}"),
        ("Total Orders Placed", str(total_orders)),
        ("Package Subscriptions", str(pkg_orders or 0)),
        ("Grocery Orders", str(frt_orders or 0)),
        ("Active Subscribers", str(active_subs or 0)),
        ("Pending Deliveries", str(pending_deliv or 0)),
    ]
    for label, val in metrics:
        ws.append([label, val])
        ws.cell(row=ws.max_row, column=1).font = regular_font
        ws.cell(row=ws.max_row, column=2).font = bold_font
    
    ws.append([])
    
    # Section 2: Top Selling Packages
    ws.append(["TOP PACKAGES PERFORMANCE", "", "", ""])
    curr_row = ws.max_row
    ws.merge_cells(f"A{curr_row}:D{curr_row}")
    ws[f"A{curr_row}"].font = section_font
    ws[f"A{curr_row}"].fill = section_fill
    ws.append(["Package Name", "Total Orders", "Revenue Generated"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.cell(row=ws.max_row, column=2).font = bold_font
    ws.cell(row=ws.max_row, column=3).font = bold_font
    
    pkg_stmt = (
        select(Product.name, func.count(Subscription.id).label('c'), func.sum(Subscription.total_amount))
        .join(Subscription, Subscription.product_id == Product.id)
        .where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed, Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED])))
        .group_by(Product.id, Product.name)
        .order_by(desc('c'))
    )
    pkg_res = await db.execute(pkg_stmt)
    for row in pkg_res.all():
        ws.append([row[0], row[1], f"Rs. {float(row[2] or 0):,.2f}"])
        
    ws.append([])
    
    # Section 3: Top Selling Groceries
    ws.append(["TOP GROCERIES PERFORMANCE", "", "", ""])
    curr_row = ws.max_row
    ws.merge_cells(f"A{curr_row}:D{curr_row}")
    ws[f"A{curr_row}"].font = section_font
    ws[f"A{curr_row}"].fill = section_fill
    ws.append(["Grocery Item", "Quantity Sold", "Revenue Generated"])
    ws.cell(row=ws.max_row, column=1).font = bold_font
    ws.cell(row=ws.max_row, column=2).font = bold_font
    ws.cell(row=ws.max_row, column=3).font = bold_font
    
    frt_stmt = (
        select(Fruit.name, func.sum(FruitOrderItem.quantity_kg).label('q'), func.sum(FruitOrderItem.price_per_kg * FruitOrderItem.quantity_kg).label('rev'))
        .join(FruitOrderItem, FruitOrderItem.fruit_id == Fruit.id)
        .join(FruitOrder, FruitOrder.id == FruitOrderItem.order_id)
        .where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed, FruitOrder.payment_status == FruitPaymentStatus.SUCCESS))
        .group_by(Fruit.id, Fruit.name)
        .order_by(desc('q'))
    )
    frt_res = await db.execute(frt_stmt)
    for row in frt_res.all():
        ws.append([row[0], int(row[1] or 0), f"Rs. {float(row[2] or 0):,.2f}"])
        
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 24
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=report_{sd}_{ed}.xlsx"},
    )

@router.get("/export/pdf")
async def export_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_super_admin)
):
    from fpdf import FPDF
    
    sd, ed = parse_dates(start_date, end_date)
    
    pkg_rev = await db.scalar(select(func.coalesce(func.sum(Payment.amount), 0)).where(and_(Payment.status == PaymentStatus.SUCCESS, func.date(Payment.paid_at) >= sd, func.date(Payment.paid_at) <= ed)))
    frt_rev = await db.scalar(select(func.coalesce(func.sum(FruitOrder.total_amount), 0)).where(and_(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS, func.date(FruitOrder.paid_at) >= sd, func.date(FruitOrder.paid_at) <= ed)))
    total_revenue = float(pkg_rev or 0) + float(frt_rev or 0)
    
    pkg_orders = await db.scalar(select(func.count(Subscription.id)).where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed, Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED]))))
    frt_orders = await db.scalar(select(func.count(FruitOrder.id)).where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed, FruitOrder.payment_status == FruitPaymentStatus.SUCCESS)))
    total_orders = (pkg_orders or 0) + (frt_orders or 0)
    
    active_subs = await db.scalar(select(func.count(Subscription.id)).where(Subscription.status == SubscriptionStatus.ACTIVE))
    pending_deliv = await db.scalar(select(func.count(SubscriptionDelivery.id)).where(and_(func.date(SubscriptionDelivery.scheduled_date) >= sd, func.date(SubscriptionDelivery.scheduled_date) <= ed, SubscriptionDelivery.status == DeliveryStatus.PENDING)))
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Header Banner (Dark Green)
    pdf.set_fill_color(27, 94, 32)
    pdf.rect(0, 0, 210, 28, 'F')
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("helvetica", size=18, style="B")
    pdf.set_xy(10, 6)
    pdf.cell(190, 8, txt="HEALTHY HOME FOODS", ln=True, align='C')
    pdf.set_font("helvetica", size=10)
    pdf.cell(190, 6, txt="Business Performance & Analytics Report", ln=True, align='C')
    
    pdf.set_text_color(60, 60, 60)
    pdf.set_xy(10, 32)
    pdf.set_font("helvetica", size=9)
    pdf.cell(95, 6, txt=f"Reporting Period: {sd.strftime('%d %b %Y')} to {ed.strftime('%d %b %Y')}", ln=False)
    pdf.cell(95, 6, txt=f"Generated On: {date.today().strftime('%d %b %Y')}", ln=True, align='R')
    pdf.ln(4)
    
    # Executive Summary Header
    pdf.set_fill_color(232, 245, 233)
    pdf.set_text_color(27, 94, 32)
    pdf.set_font("helvetica", size=12, style="B")
    pdf.cell(190, 8, txt="  Executive Overview", ln=True, fill=True)
    pdf.ln(2)
    
    pdf.set_text_color(40, 40, 40)
    pdf.set_font("helvetica", size=10)
    
    metrics = [
        ("Total Revenue", f"Rs. {total_revenue:,.2f}", "Total Orders", str(total_orders)),
        ("Package Revenue", f"Rs. {float(pkg_rev or 0):,.2f}", "Package Subscriptions", str(pkg_orders or 0)),
        ("Grocery Revenue", f"Rs. {float(frt_rev or 0):,.2f}", "Grocery Orders", str(frt_orders or 0)),
        ("Active Subscribers", str(active_subs or 0), "Pending Deliveries", str(pending_deliv or 0)),
    ]
    for m1_l, m1_v, m2_l, m2_v in metrics:
        pdf.set_font("helvetica", style="", size=9)
        pdf.cell(45, 7, txt=m1_l, border='B')
        pdf.set_font("helvetica", style="B", size=9)
        pdf.cell(50, 7, txt=m1_v, border='B')
        pdf.set_font("helvetica", style="", size=9)
        pdf.cell(45, 7, txt=m2_l, border='B')
        pdf.set_font("helvetica", style="B", size=9)
        pdf.cell(50, 7, txt=m2_v, border='B', ln=True)
        
    pdf.ln(6)
    
    # Top Selling Packages
    pdf.set_fill_color(232, 245, 233)
    pdf.set_text_color(27, 94, 32)
    pdf.set_font("helvetica", size=12, style="B")
    pdf.cell(190, 8, txt="  Top Packages Performance", ln=True, fill=True)
    pdf.ln(2)
    
    pdf.set_fill_color(245, 245, 245)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", style="B", size=9)
    pdf.cell(90, 7, txt="Package Name", border=1, fill=True)
    pdf.cell(40, 7, txt="Orders", border=1, align='C', fill=True)
    pdf.cell(60, 7, txt="Revenue", border=1, align='R', fill=True, ln=True)
    
    pkg_stmt = (
        select(Product.name, func.count(Subscription.id).label('c'), func.sum(Subscription.total_amount))
        .join(Subscription, Subscription.product_id == Product.id)
        .where(and_(func.date(Subscription.created_at) >= sd, func.date(Subscription.created_at) <= ed, Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED])))
        .group_by(Product.id, Product.name)
        .order_by(desc('c'))
        .limit(6)
    )
    pkg_res = await db.execute(pkg_stmt)
    pdf.set_font("helvetica", size=9)
    for row in pkg_res.all():
        pdf.cell(90, 6, txt=str(row[0]), border=1)
        pdf.cell(40, 6, txt=str(row[1]), border=1, align='C')
        pdf.cell(60, 6, txt=f"Rs. {float(row[2] or 0):,.2f}", border=1, align='R', ln=True)
        
    pdf.ln(6)
    
    # Top Selling Groceries
    pdf.set_fill_color(232, 245, 233)
    pdf.set_text_color(27, 94, 32)
    pdf.set_font("helvetica", size=12, style="B")
    pdf.cell(190, 8, txt="  Top Groceries Performance", ln=True, fill=True)
    pdf.ln(2)
    
    pdf.set_fill_color(245, 245, 245)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", style="B", size=9)
    pdf.cell(90, 7, txt="Grocery Item", border=1, fill=True)
    pdf.cell(40, 7, txt="Quantity Sold", border=1, align='C', fill=True)
    pdf.cell(60, 7, txt="Revenue", border=1, align='R', fill=True, ln=True)
    
    frt_stmt = (
        select(Fruit.name, func.sum(FruitOrderItem.quantity_kg).label('q'), func.sum(FruitOrderItem.price_per_kg * FruitOrderItem.quantity_kg).label('rev'))
        .join(FruitOrderItem, FruitOrderItem.fruit_id == Fruit.id)
        .join(FruitOrder, FruitOrder.id == FruitOrderItem.order_id)
        .where(and_(func.date(FruitOrder.created_at) >= sd, func.date(FruitOrder.created_at) <= ed, FruitOrder.payment_status == FruitPaymentStatus.SUCCESS))
        .group_by(Fruit.id, Fruit.name)
        .order_by(desc('q'))
        .limit(6)
    )
    frt_res = await db.execute(frt_stmt)
    pdf.set_font("helvetica", size=9)
    for row in frt_res.all():
        pdf.cell(90, 6, txt=str(row[0]), border=1)
        pdf.cell(40, 6, txt=str(int(row[1] or 0)), border=1, align='C')
        pdf.cell(60, 6, txt=f"Rs. {float(row[2] or 0):,.2f}", border=1, align='R', ln=True)
        
    pdf_bytes = bytes(pdf.output())
    
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename=report_{sd}_{ed}.pdf"}
    )

@router.get("/admin-overview")
async def get_admin_overview(
    _: User = Depends(require_super_admin),
    db: AsyncSession = Depends(get_db),
):
    today = date.today()
    sd_7 = today - timedelta(days=6)
    
    # Summary
    todays_pkg_rev = await db.scalar(select(func.sum(Payment.amount)).where(and_(Payment.status == PaymentStatus.SUCCESS, func.date(Payment.paid_at) == today)))
    todays_frt_rev = await db.scalar(select(func.sum(FruitOrder.total_amount)).where(and_(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS, func.date(FruitOrder.paid_at) == today)))
    todays_revenue = float(todays_pkg_rev or 0) + float(todays_frt_rev or 0)
    
    orders_pkg = await db.scalar(
        select(func.count(Subscription.id.distinct()))
        .join(Payment, Payment.subscription_id == Subscription.id)
        .where(
            and_(
                Payment.status == PaymentStatus.SUCCESS,
                func.date(Payment.paid_at) == today
            )
        )
    )
    orders_frt = await db.scalar(
        select(func.count(FruitOrder.id))
        .where(
            and_(
                func.date(FruitOrder.paid_at) == today,
                FruitOrder.payment_status == FruitPaymentStatus.SUCCESS
            )
        )
    )
    orders_today = (orders_pkg or 0) + (orders_frt or 0)
    
    active_subscribers = await db.scalar(select(func.count(Subscription.id)).where(Subscription.status == SubscriptionStatus.ACTIVE))
    
    pending_deliveries = await db.scalar(select(func.count(SubscriptionDelivery.id)).where(and_(SubscriptionDelivery.status == DeliveryStatus.PENDING, func.date(SubscriptionDelivery.scheduled_date) == today)))
    
    # Insights
    top_pkg_res = await db.execute(
        select(Product.name, func.count(Subscription.id).label('c'))
        .join(Subscription, Subscription.product_id == Product.id)
        .where(Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED]))
        .group_by(Product.id)
        .order_by(desc('c'))
        .limit(1)
    )
    top_pkg_row = top_pkg_res.first()
    top_selling_package = {"name": top_pkg_row[0], "count": top_pkg_row[1]} if top_pkg_row else None
    
    top_frt_res = await db.execute(
        select(Fruit.name, func.count(FruitOrderItem.id).label('c'))
        .join(FruitOrderItem, FruitOrderItem.fruit_id == Fruit.id)
        .join(FruitOrder, FruitOrder.id == FruitOrderItem.order_id)
        .where(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS)
        .group_by(Fruit.id)
        .order_by(desc('c'))
        .limit(1)
    )
    top_frt_row = top_frt_res.first()
    most_ordered_fruit = {"name": top_frt_row[0], "count": top_frt_row[1]} if top_frt_row else None
    
    # Revenue Chart (Last 7 days)
    chart_data = []
    for i in range(7):
        d = sd_7 + timedelta(days=i)
        p_rev = await db.scalar(select(func.sum(Payment.amount)).where(and_(Payment.status == PaymentStatus.SUCCESS, func.date(Payment.paid_at) == d)))
        f_rev = await db.scalar(select(func.sum(FruitOrder.total_amount)).where(and_(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS, func.date(FruitOrder.paid_at) == d)))
        chart_data.append({
            "date": d.strftime("%Y-%m-%d"),
            "revenue": float(p_rev or 0) + float(f_rev or 0)
        })
        
    # Recent Activity Logic
    activities = []
    
    # 1. Fruit Orders
    frt_stmt = select(FruitOrder, User.full_name).join(Customer, FruitOrder.customer_id == Customer.id).join(User, Customer.user_id == User.id).where(FruitOrder.payment_status == FruitPaymentStatus.SUCCESS).order_by(FruitOrder.created_at.desc()).limit(5)
    frt_res = await db.execute(frt_stmt)
    for fo, fname in frt_res:
        activities.append({
            "type": "fruit_order", "title": "New Fruit Order",
            "description": f"{fname or 'Customer'} placed Order #{fo.order_number}",
            "timestamp": fo.created_at
        })

    # 2. Subscriptions
    sub_stmt = select(Subscription, User.full_name).join(Customer, Subscription.customer_id == Customer.id).join(User, Customer.user_id == User.id).where(Subscription.status.in_([SubscriptionStatus.ACTIVE, SubscriptionStatus.PAUSED, SubscriptionStatus.COMPLETED])).order_by(Subscription.created_at.desc()).limit(5)
    sub_res = await db.execute(sub_stmt)
    for sub, fname in sub_res:
        activities.append({
            "type": "subscription", "title": "New Subscription",
            "description": f"{fname or 'Customer'} purchased a package",
            "timestamp": sub.created_at
        })

    # 3. Deliveries
    del_stmt = select(DeliveryAssignment, User.full_name).join(DeliveryPartner, DeliveryAssignment.delivery_partner_id == DeliveryPartner.id).join(User, DeliveryPartner.user_id == User.id).order_by(func.coalesce(DeliveryAssignment.delivered_at, DeliveryAssignment.out_at, DeliveryAssignment.assigned_at).desc()).limit(5)
    del_res = await db.execute(del_stmt)
    for assign, dpname in del_res:
        ts = assign.delivered_at or assign.out_at or assign.assigned_at
        if assign.status == AssignmentStatus.DELIVERED:
            title, activity_desc = "Order Delivered", f"{dpname} delivered an order"
        elif assign.status == AssignmentStatus.OUT_FOR_DELIVERY:
            title, activity_desc = "Out for Delivery", f"{dpname} is out for delivery"
        else:
            title, activity_desc = "Delivery Partner Assigned", f"Assigned to {dpname}"
        activities.append({
            "type": "delivery", "title": title, "description": activity_desc, "timestamp": ts
        })

    # 4. Customers
    cust_stmt = select(Customer, User.full_name).join(User, Customer.user_id == User.id).order_by(Customer.created_at.desc()).limit(5)
    cust_res = await db.execute(cust_stmt)
    for cust, fname in cust_res:
        activities.append({
            "type": "customer", "title": "New Customer",
            "description": f"{fname or 'A new customer'} registered",
            "timestamp": cust.created_at
        })

    activities.sort(key=lambda x: x["timestamp"] if x["timestamp"] else datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    recent_activity = []
    for a in activities[:5]:
        a["timestamp"] = a["timestamp"].isoformat() if a["timestamp"] else None
        recent_activity.append(a)

    return {
        "summary": {
            "todays_revenue": todays_revenue,
            "orders_today": orders_today,
            "active_subscribers": active_subscribers or 0,
            "pending_deliveries": pending_deliveries or 0,
        },
        "quick_insights": {
            "top_selling_package": top_selling_package,
            "most_ordered_fruit": most_ordered_fruit,
        },
        "revenue_chart": chart_data,
        "recent_activity": recent_activity
    }
